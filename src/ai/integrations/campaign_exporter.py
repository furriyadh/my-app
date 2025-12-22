#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
📤 Campaign Exporter - مُصدر الحملات
====================================

مُصدر شامل للحملات يدعم تصدير البيانات بصيغ متعددة:
- JSON, CSV, Excel, XML
- تقارير مخصصة
- تصدير جماعي
- ضغط الملفات
- دعم MCC

المطور: Google Ads AI Platform Team
التاريخ: 2025-07-07
الإصدار: 1.0.0
"""

import logging
import json
import csv
import xml.etree.ElementTree as ET
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
from dataclasses import dataclass, field
from enum import Enum
import os
import zipfile
import io
import base64

# استيراد المكتبات الاختيارية
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    Workbook = None

# استيراد وحدات النظام
try:
    from ..utils.logger import setup_logger
    logger = setup_logger(__name__)
except ImportError:
    logger = logging.getLogger(__name__)

class ExportFormat(Enum):
    """صيغ التصدير المدعومة"""
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    XML = "xml"
    TXT = "txt"
    HTML = "html"

class CompressionType(Enum):
    """أنواع الضغط"""
    NONE = "none"
    ZIP = "zip"
    GZIP = "gzip"

@dataclass
class ExportConfig:
    """
    ⚙️ إعدادات التصدير
    """
    format: ExportFormat = ExportFormat.JSON
    compression: CompressionType = CompressionType.NONE
    include_metadata: bool = True
    include_statistics: bool = True
    date_format: str = "%Y-%m-%d %H:%M:%S"
    encoding: str = "utf-8"
    pretty_print: bool = True
    custom_fields: List[str] = field(default_factory=list)
    exclude_fields: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        """تحويل إلى قاموس"""
        return {
            'format': self.format.value,
            'compression': self.compression.value,
            'include_metadata': self.include_metadata,
            'include_statistics': self.include_statistics,
            'date_format': self.date_format,
            'encoding': self.encoding,
            'pretty_print': self.pretty_print,
            'custom_fields': self.custom_fields,
            'exclude_fields': self.exclude_fields
        }

@dataclass
class ExportResult:
    """
    📊 نتيجة التصدير
    """
    success: bool
    file_path: Optional[str] = None
    file_size: int = 0
    records_count: int = 0
    format: Optional[ExportFormat] = None
    compression: Optional[CompressionType] = None
    export_time: float = 0.0
    error_message: str = ""
    warnings: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        """تحويل إلى قاموس"""
        return {
            'success': self.success,
            'file_path': self.file_path,
            'file_size': self.file_size,
            'records_count': self.records_count,
            'format': self.format.value if self.format else None,
            'compression': self.compression.value if self.compression else None,
            'export_time': self.export_time,
            'error_message': self.error_message,
            'warnings': self.warnings,
            'metadata': self.metadata
        }

class CampaignExporter:
    """
    📤 مُصدر الحملات المتقدم
    
    يوفر تصدير شامل للحملات والبيانات المرتبطة بها
    بصيغ متعددة مع دعم الضغط والتخصيص.
    """
    
    def __init__(self, format_type: str = "json", output_dir: str = "/tmp"):
        """
        تهيئة المُصدر
        
        Args:
            format_type: نوع الصيغة الافتراضية
            output_dir: مجلد الإخراج
        """
        self.default_format = ExportFormat(format_type.lower())
        self.output_dir = output_dir
        
        # إنشاء مجلد الإخراج إذا لم يكن موجوداً
        os.makedirs(output_dir, exist_ok=True)
        
        # إحصائيات التصدير
        self.export_stats = {
            'total_exports': 0,
            'successful_exports': 0,
            'failed_exports': 0,
            'total_records_exported': 0,
            'total_file_size': 0
        }
        
        logger.info(f"📤 تم تهيئة مُصدر الحملات - الصيغة الافتراضية: {self.default_format.value}")
    
    async def export_campaigns(
        self,
        campaigns_data: List[Dict[str, Any]],
        config: Optional[ExportConfig] = None,
        filename: Optional[str] = None
    ) -> ExportResult:
        """
        تصدير بيانات الحملات
        
        Args:
            campaigns_data: بيانات الحملات
            config: إعدادات التصدير
            filename: اسم الملف (اختياري)
            
        Returns:
            ExportResult: نتيجة التصدير
        """
        start_time = datetime.now()
        
        # استخدام الإعدادات الافتراضية إذا لم تُحدد
        if config is None:
            config = ExportConfig(format=self.default_format)
        
        # إنشاء اسم الملف إذا لم يُحدد
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"campaigns_export_{timestamp}"
        
        try:
            # تحضير البيانات للتصدير
            export_data = await self._prepare_export_data(campaigns_data, config)
            
            # تصدير حسب الصيغة
            if config.format == ExportFormat.JSON:
                result = await self._export_json(export_data, config, filename)
            elif config.format == ExportFormat.CSV:
                result = await self._export_csv(export_data, config, filename)
            elif config.format == ExportFormat.EXCEL:
                result = await self._export_excel(export_data, config, filename)
            elif config.format == ExportFormat.XML:
                result = await self._export_xml(export_data, config, filename)
            elif config.format == ExportFormat.TXT:
                result = await self._export_txt(export_data, config, filename)
            elif config.format == ExportFormat.HTML:
                result = await self._export_html(export_data, config, filename)
            else:
                raise ValueError(f"صيغة غير مدعومة: {config.format}")
            
            # تطبيق الضغط إذا طُلب
            if config.compression != CompressionType.NONE and result.success:
                result = await self._apply_compression(result, config)
            
            # حساب وقت التصدير
            export_time = (datetime.now() - start_time).total_seconds()
            result.export_time = export_time
            
            # تحديث الإحصائيات
            self.export_stats['total_exports'] += 1
            if result.success:
                self.export_stats['successful_exports'] += 1
                self.export_stats['total_records_exported'] += result.records_count
                self.export_stats['total_file_size'] += result.file_size
            else:
                self.export_stats['failed_exports'] += 1
            
            logger.info(f"📤 {'نجح' if result.success else 'فشل'} تصدير {len(campaigns_data)} حملة في {export_time:.2f}s")
            
            return result
            
        except Exception as e:
            error_result = ExportResult(
                success=False,
                error_message=str(e),
                format=config.format,
                compression=config.compression,
                export_time=(datetime.now() - start_time).total_seconds()
            )
            
            self.export_stats['total_exports'] += 1
            self.export_stats['failed_exports'] += 1
            
            logger.error(f"❌ فشل في تصدير الحملات: {e}")
            return error_result
    
    async def export_multiple_formats(
        self,
        campaigns_data: List[Dict[str, Any]],
        formats: List[ExportFormat],
        base_filename: Optional[str] = None
    ) -> List[ExportResult]:
        """
        تصدير بصيغ متعددة
        
        Args:
            campaigns_data: بيانات الحملات
            formats: قائمة الصيغ
            base_filename: اسم الملف الأساسي
            
        Returns:
            List[ExportResult]: نتائج التصدير
        """
        if base_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"campaigns_multi_export_{timestamp}"
        
        results = []
        
        for format_type in formats:
            config = ExportConfig(format=format_type)
            filename = f"{base_filename}_{format_type.value}"
            
            result = await self.export_campaigns(campaigns_data, config, filename)
            results.append(result)
        
        successful = sum(1 for r in results if r.success)
        logger.info(f"📦 تصدير متعدد الصيغ: {successful}/{len(formats)} نجح")
        
        return results
    
    async def _prepare_export_data(
        self,
        campaigns_data: List[Dict[str, Any]],
        config: ExportConfig
    ) -> Dict[str, Any]:
        """تحضير البيانات للتصدير"""
        
        # تصفية الحقول المطلوبة
        filtered_data = []
        for campaign in campaigns_data:
            filtered_campaign = {}
            
            for key, value in campaign.items():
                # تخطي الحقول المستبعدة
                if key in config.exclude_fields:
                    continue
                
                # تضمين الحقول المخصصة فقط إذا حُددت
                if config.custom_fields and key not in config.custom_fields:
                    continue
                
                # تنسيق التواريخ
                if isinstance(value, datetime):
                    value = value.strftime(config.date_format)
                
                filtered_campaign[key] = value
            
            filtered_data.append(filtered_campaign)
        
        # إعداد البيانات النهائية
        export_data = {
            'campaigns': filtered_data,
            'total_count': len(filtered_data)
        }
        
        # إضافة البيانات الوصفية
        if config.include_metadata:
            export_data['metadata'] = {
                'export_timestamp': datetime.now().strftime(config.date_format),
                'export_format': config.format.value,
                'total_campaigns': len(filtered_data),
                'exporter_version': '1.0.0'
            }
        
        # إضافة الإحصائيات
        if config.include_statistics:
            export_data['statistics'] = await self._calculate_statistics(filtered_data)
        
        return export_data
    
    async def _calculate_statistics(self, campaigns_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """حساب إحصائيات البيانات"""
        if not campaigns_data:
            return {}
        
        stats = {
            'total_campaigns': len(campaigns_data),
            'active_campaigns': 0,
            'paused_campaigns': 0,
            'total_budget': 0.0,
            'total_impressions': 0,
            'total_clicks': 0,
            'total_cost': 0.0,
            'average_ctr': 0.0,
            'average_cpc': 0.0
        }
        
        total_impressions = 0
        total_clicks = 0
        total_cost = 0.0
        
        for campaign in campaigns_data:
            # حالة الحملة
            status = campaign.get('status', '').upper()
            if status == 'ENABLED':
                stats['active_campaigns'] += 1
            elif status == 'PAUSED':
                stats['paused_campaigns'] += 1
            
            # الميزانية
            budget = float(campaign.get('budget', 0))
            stats['total_budget'] += budget
            
            # المقاييس
            impressions = int(campaign.get('impressions', 0))
            clicks = int(campaign.get('clicks', 0))
            cost = float(campaign.get('cost', 0))
            
            total_impressions += impressions
            total_clicks += clicks
            total_cost += cost
        
        stats['total_impressions'] = total_impressions
        stats['total_clicks'] = total_clicks
        stats['total_cost'] = total_cost
        
        # حساب المتوسطات
        if total_impressions > 0:
            stats['average_ctr'] = round((total_clicks / total_impressions) * 100, 2)
        
        if total_clicks > 0:
            stats['average_cpc'] = round(total_cost / total_clicks, 2)
        
        return stats
    
    async def _export_json(
        self,
        data: Dict[str, Any],
        config: ExportConfig,
        filename: str
    ) -> ExportResult:
        """تصدير بصيغة JSON"""
        
        file_path = os.path.join(self.output_dir, f"{filename}.json")
        
        try:
            with open(file_path, 'w', encoding=config.encoding) as f:
                if config.pretty_print:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                else:
                    json.dump(data, f, ensure_ascii=False)
            
            file_size = os.path.getsize(file_path)
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                records_count=data.get('total_count', 0),
                format=ExportFormat.JSON,
                compression=config.compression
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"فشل في تصدير JSON: {e}",
                format=ExportFormat.JSON
            )
    
    async def _export_csv(
        self,
        data: Dict[str, Any],
        config: ExportConfig,
        filename: str
    ) -> ExportResult:
        """تصدير بصيغة CSV"""
        
        file_path = os.path.join(self.output_dir, f"{filename}.csv")
        campaigns = data.get('campaigns', [])
        
        if not campaigns:
            return ExportResult(
                success=False,
                error_message="لا توجد بيانات للتصدير",
                format=ExportFormat.CSV
            )
        
        try:
            with open(file_path, 'w', newline='', encoding=config.encoding) as f:
                # الحصول على أسماء الأعمدة
                fieldnames = list(campaigns[0].keys())
                
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(campaigns)
            
            file_size = os.path.getsize(file_path)
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                records_count=len(campaigns),
                format=ExportFormat.CSV,
                compression=config.compression
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"فشل في تصدير CSV: {e}",
                format=ExportFormat.CSV
            )
    
    async def _export_excel(
        self,
        data: Dict[str, Any],
        config: ExportConfig,
        filename: str
    ) -> ExportResult:
        """تصدير بصيغة Excel"""
        
        if not EXCEL_AVAILABLE:
            return ExportResult(
                success=False,
                error_message="مكتبة openpyxl غير متاحة",
                format=ExportFormat.EXCEL
            )
        
        file_path = os.path.join(self.output_dir, f"{filename}.xlsx")
        campaigns = data.get('campaigns', [])
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Campaigns"
            
            if campaigns:
                # كتابة العناوين
                headers = list(campaigns[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # كتابة البيانات
                for row, campaign in enumerate(campaigns, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=campaign.get(header, ''))
            
            # إضافة ورقة الإحصائيات
            if config.include_statistics and 'statistics' in data:
                stats_ws = wb.create_sheet("Statistics")
                stats = data['statistics']
                
                for row, (key, value) in enumerate(stats.items(), 1):
                    stats_ws.cell(row=row, column=1, value=key)
                    stats_ws.cell(row=row, column=2, value=value)
            
            wb.save(file_path)
            file_size = os.path.getsize(file_path)
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                records_count=len(campaigns),
                format=ExportFormat.EXCEL,
                compression=config.compression
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"فشل في تصدير Excel: {e}",
                format=ExportFormat.EXCEL
            )
    
    async def _export_xml(
        self,
        data: Dict[str, Any],
        config: ExportConfig,
        filename: str
    ) -> ExportResult:
        """تصدير بصيغة XML"""
        
        file_path = os.path.join(self.output_dir, f"{filename}.xml")
        
        try:
            root = ET.Element("campaigns_export")
            
            # إضافة البيانات الوصفية
            if config.include_metadata and 'metadata' in data:
                metadata_elem = ET.SubElement(root, "metadata")
                for key, value in data['metadata'].items():
                    elem = ET.SubElement(metadata_elem, key)
                    elem.text = str(value)
            
            # إضافة الحملات
            campaigns_elem = ET.SubElement(root, "campaigns")
            for campaign in data.get('campaigns', []):
                campaign_elem = ET.SubElement(campaigns_elem, "campaign")
                for key, value in campaign.items():
                    elem = ET.SubElement(campaign_elem, key)
                    elem.text = str(value) if value is not None else ""
            
            # إضافة الإحصائيات
            if config.include_statistics and 'statistics' in data:
                stats_elem = ET.SubElement(root, "statistics")
                for key, value in data['statistics'].items():
                    elem = ET.SubElement(stats_elem, key)
                    elem.text = str(value)
            
            # كتابة الملف
            tree = ET.ElementTree(root)
            tree.write(file_path, encoding=config.encoding, xml_declaration=True)
            
            file_size = os.path.getsize(file_path)
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                records_count=data.get('total_count', 0),
                format=ExportFormat.XML,
                compression=config.compression
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"فشل في تصدير XML: {e}",
                format=ExportFormat.XML
            )
    
    async def _export_txt(
        self,
        data: Dict[str, Any],
        config: ExportConfig,
        filename: str
    ) -> ExportResult:
        """تصدير بصيغة TXT"""
        
        file_path = os.path.join(self.output_dir, f"{filename}.txt")
        
        try:
            with open(file_path, 'w', encoding=config.encoding) as f:
                # كتابة العنوان
                f.write("تقرير تصدير الحملات\n")
                f.write("=" * 50 + "\n\n")
                
                # كتابة البيانات الوصفية
                if config.include_metadata and 'metadata' in data:
                    f.write("البيانات الوصفية:\n")
                    f.write("-" * 20 + "\n")
                    for key, value in data['metadata'].items():
                        f.write(f"{key}: {value}\n")
                    f.write("\n")
                
                # كتابة الإحصائيات
                if config.include_statistics and 'statistics' in data:
                    f.write("الإحصائيات:\n")
                    f.write("-" * 20 + "\n")
                    for key, value in data['statistics'].items():
                        f.write(f"{key}: {value}\n")
                    f.write("\n")
                
                # كتابة الحملات
                f.write("الحملات:\n")
                f.write("-" * 20 + "\n")
                for i, campaign in enumerate(data.get('campaigns', []), 1):
                    f.write(f"\nحملة {i}:\n")
                    for key, value in campaign.items():
                        f.write(f"  {key}: {value}\n")
            
            file_size = os.path.getsize(file_path)
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                records_count=data.get('total_count', 0),
                format=ExportFormat.TXT,
                compression=config.compression
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"فشل في تصدير TXT: {e}",
                format=ExportFormat.TXT
            )
    
    async def _export_html(
        self,
        data: Dict[str, Any],
        config: ExportConfig,
        filename: str
    ) -> ExportResult:
        """تصدير بصيغة HTML"""
        
        file_path = os.path.join(self.output_dir, f"{filename}.html")
        campaigns = data.get('campaigns', [])
        
        try:
            html_content = """
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>تقرير تصدير الحملات</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        table { border-collapse: collapse; width: 100%; margin: 20px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: right; }
        th { background-color: #f2f2f2; }
        .metadata, .statistics { background-color: #f9f9f9; padding: 15px; margin: 10px 0; }
        h1, h2 { color: #333; }
    </style>
</head>
<body>
    <h1>تقرير تصدير الحملات</h1>
"""
            
            # إضافة البيانات الوصفية
            if config.include_metadata and 'metadata' in data:
                html_content += "<div class='metadata'><h2>البيانات الوصفية</h2>"
                for key, value in data['metadata'].items():
                    html_content += f"<p><strong>{key}:</strong> {value}</p>"
                html_content += "</div>"
            
            # إضافة الإحصائيات
            if config.include_statistics and 'statistics' in data:
                html_content += "<div class='statistics'><h2>الإحصائيات</h2>"
                for key, value in data['statistics'].items():
                    html_content += f"<p><strong>{key}:</strong> {value}</p>"
                html_content += "</div>"
            
            # إضافة جدول الحملات
            if campaigns:
                html_content += "<h2>الحملات</h2><table>"
                
                # العناوين
                headers = list(campaigns[0].keys())
                html_content += "<tr>"
                for header in headers:
                    html_content += f"<th>{header}</th>"
                html_content += "</tr>"
                
                # البيانات
                for campaign in campaigns:
                    html_content += "<tr>"
                    for header in headers:
                        value = campaign.get(header, '')
                        html_content += f"<td>{value}</td>"
                    html_content += "</tr>"
                
                html_content += "</table>"
            
            html_content += "</body></html>"
            
            with open(file_path, 'w', encoding=config.encoding) as f:
                f.write(html_content)
            
            file_size = os.path.getsize(file_path)
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                records_count=len(campaigns),
                format=ExportFormat.HTML,
                compression=config.compression
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"فشل في تصدير HTML: {e}",
                format=ExportFormat.HTML
            )
    
    async def _apply_compression(
        self,
        result: ExportResult,
        config: ExportConfig
    ) -> ExportResult:
        """تطبيق الضغط على الملف"""
        
        if not result.success or not result.file_path:
            return result
        
        try:
            if config.compression == CompressionType.ZIP:
                # ضغط ZIP
                zip_path = result.file_path + ".zip"
                
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    zipf.write(result.file_path, os.path.basename(result.file_path))
                
                # حذف الملف الأصلي
                os.remove(result.file_path)
                
                # تحديث النتيجة
                result.file_path = zip_path
                result.file_size = os.path.getsize(zip_path)
                result.compression = CompressionType.ZIP
                
                logger.debug(f"✅ تم ضغط الملف: {zip_path}")
            
            return result
            
        except Exception as e:
            result.warnings.append(f"فشل في ضغط الملف: {e}")
            logger.warning(f"⚠️ فشل في ضغط الملف: {e}")
            return result
    
    def get_export_stats(self) -> Dict[str, Any]:
        """الحصول على إحصائيات التصدير"""
        stats = self.export_stats.copy()
        
        if stats['total_exports'] > 0:
            stats['success_rate'] = (stats['successful_exports'] / stats['total_exports']) * 100
        else:
            stats['success_rate'] = 0
        
        return stats
    
    def reset_export_stats(self):
        """إعادة تعيين إحصائيات التصدير"""
        self.export_stats = {
            'total_exports': 0,
            'successful_exports': 0,
            'failed_exports': 0,
            'total_records_exported': 0,
            'total_file_size': 0
        }
        logger.info("📊 تم إعادة تعيين إحصائيات التصدير")

# دوال مساعدة للاستخدام السريع
def get_campaign_exporter(format_type: str = "json", output_dir: str = "/tmp") -> CampaignExporter:
    """الحصول على مُصدر الحملات"""
    return CampaignExporter(format_type=format_type, output_dir=output_dir)

async def export_campaigns_quick(
    campaigns_data: List[Dict[str, Any]],
    format_type: str = "json",
    filename: Optional[str] = None,
    output_dir: str = "/tmp"
) -> ExportResult:
    """تصدير سريع للحملات"""
    exporter = get_campaign_exporter(format_type, output_dir)
    config = ExportConfig(format=ExportFormat(format_type.lower()))
    return await exporter.export_campaigns(campaigns_data, config, filename)

# تصدير الوحدات المهمة
__all__ = [
    'CampaignExporter',
    'ExportConfig',
    'ExportResult',
    'ExportFormat',
    'CompressionType',
    'get_campaign_exporter',
    'export_campaigns_quick'
]

